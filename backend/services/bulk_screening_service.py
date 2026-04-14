"""Bulk screening service: CSV parsing, batch screening, Excel generation."""
import csv
import io
import uuid
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CSV_TEMPLATE_ROWS = [
    ["name", "dob", "nationality", "id_type", "id_number"],
    ["Rajesh Kumar Sharma", "1985-03-15", "IN", "PAN", "ABCPS1234D"],
    ["Ananya Textiles Pvt Ltd", "", "IN", "", ""],
    ["Deepak Malhotra", "1978-11-22", "IN", "AADHAAR", "987654321012"],
]


def generate_csv_template() -> str:
    """Generate a CSV template with headers and 3 example rows."""
    output = io.StringIO()
    writer = csv.writer(output)
    for row in CSV_TEMPLATE_ROWS:
        writer.writerow(row)
    return output.getvalue()


def parse_csv(content: str) -> list[dict]:
    """Parse CSV content into a list of entity dicts. Returns parsed rows."""
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for i, row in enumerate(reader):
        name = (row.get("name") or "").strip()
        if not name:
            continue
        rows.append({
            "row_num": i + 2,
            "name": name,
            "dob": (row.get("dob") or "").strip() or None,
            "nationality": (row.get("nationality") or "").strip() or None,
            "id_type": (row.get("id_type") or "").strip() or None,
            "id_number": (row.get("id_number") or "").strip() or None,
        })
    return rows


def generate_results_excel(batch: dict, results: list[dict]) -> bytes:
    """Generate a branded Excel file with Summary + Detailed Results sheets."""
    wb = Workbook()

    # Colors
    header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    accent_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    brand_font = Font(color="2563EB", bold=True, size=14)
    sub_font = Font(color="94A3B8", size=10)
    thin_border = Border(
        left=Side(style="thin", color="1E2530"),
        right=Side(style="thin", color="1E2530"),
        top=Side(style="thin", color="1E2530"),
        bottom=Side(style="thin", color="1E2530"),
    )

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = "2563EB"

    # Branding header
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "Rudrik.io — Bulk Screening Report"
    ws1["A1"].font = brand_font
    ws1["A1"].alignment = Alignment(vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:F2")
    ws1["A2"] = "Compliance Intelligence Platform"
    ws1["A2"].font = sub_font

    # Summary data
    total = len(results)
    matches = sum(1 for r in results if r.get("has_match"))
    high_risk = sum(1 for r in results if r.get("risk_level") in ("HIGH", "CRITICAL"))
    medium_risk = sum(1 for r in results if r.get("risk_level") == "MEDIUM")
    low_risk = sum(1 for r in results if r.get("risk_level") == "LOW")

    summary_rows = [
        ("Batch ID", batch.get("batch_id", "")),
        ("Screening Date", batch.get("created_at", "")),
        ("Screening Mode", batch.get("mode", "demo").upper()),
        ("Total Entities Screened", total),
        ("Total Matches Found", matches),
        ("High/Critical Risk", high_risk),
        ("Medium Risk", medium_risk),
        ("Low Risk", low_risk),
        ("Match Rate", f"{round(matches / total * 100, 1)}%" if total > 0 else "0%"),
    ]

    for i, (label, value) in enumerate(summary_rows, start=4):
        ws1[f"A{i}"] = label
        ws1[f"A{i}"].font = Font(bold=True, size=10, color="F1F5F9")
        ws1[f"A{i}"].fill = header_fill
        ws1[f"B{i}"] = value
        ws1[f"B{i}"].font = Font(size=10, color="F1F5F9")
        ws1[f"B{i}"].fill = header_fill
        ws1[f"A{i}"].border = thin_border
        ws1[f"B{i}"].border = thin_border

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    # --- Sheet 2: Detailed Results ---
    ws2 = wb.create_sheet("Detailed Results")
    ws2.sheet_properties.tabColor = "FF6B35"

    # Branding
    ws2.merge_cells("A1:J1")
    ws2["A1"] = "Rudrik.io — Detailed Screening Results"
    ws2["A1"].font = brand_font
    ws2.row_dimensions[1].height = 28

    headers = [
        "Name", "DOB", "Nationality", "ID Type", "ID Number",
        "Risk Score", "Risk Level", "Match Found", "Match Types", "SLA Status",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, r in enumerate(results, start=4):
        match_types = []
        if r.get("sanctions_match"):
            match_types.append("Sanction")
        if r.get("pep_match"):
            match_types.append("PEP")
        if r.get("adverse_media_match"):
            match_types.append("Adverse Media")

        row_data = [
            r.get("full_name", ""),
            r.get("date_of_birth", "") or "",
            r.get("nationality", "") or "",
            r.get("id_type", "") or "",
            r.get("id_number", "") or "",
            r.get("risk_score", 0),
            r.get("risk_level", "LOW"),
            "Yes" if r.get("has_match") else "No",
            ", ".join(match_types) if match_types else "None",
            r.get("sla_status", "on_time").replace("_", " ").title(),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(size=10, color="F1F5F9")
            cell.fill = header_fill
            cell.border = thin_border

            # Color coding for risk level
            if col == 7:
                color_map = {"LOW": "10B981", "MEDIUM": "F59E0B", "HIGH": "EF4444", "CRITICAL": "DC2626"}
                cell.font = Font(size=10, bold=True, color=color_map.get(str(val), "F1F5F9"))
            # Color coding for match found
            if col == 8:
                cell.font = Font(size=10, bold=True, color="EF4444" if val == "Yes" else "10B981")

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 4)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["I"].width = 24

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
